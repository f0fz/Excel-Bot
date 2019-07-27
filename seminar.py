#!/usr/bin/env python
# -*- coding: utf-8 -*-

from telegram import ReplyKeyboardMarkup
from telegram.ext import (Updater, CommandHandler, MessageHandler, Filters, RegexHandler,
                          ConversationHandler)

from openpyxl import load_workbook
from excel import returnSeating, saveFile

import logging, threading

logging.basicConfig(format='%(asctime)s - %(name)s - %(levelname)s - %(message)s',
                    level=logging.INFO)

logger = logging.getLogger(__name__)

###############
### <EXCEL> ###

# read working copy of the workbook
main_workbook = load_workbook('SeminarDatasheet.xlsx')
ws = main_workbook['Sheet1']
logger.info("Opening Excel list...")
# dump all info values into PERSON dict array for access
PERSON = []
row_number = 2
while ws['A' + str(row_number)].value != None:
    PERSON.append({'NRIC': ws['A' + str(row_number)].value,
                   'GRP1': ws['B' + str(row_number)].value,
                   'GRP1_REG': ''})
    row_number += 1
logger.info("Excel list dumped in memory")

### </EXCEL> ###
################

# init values for Telegram
TYPING_NRIC, RESPONSE = range(2)
reply_keyboard = [['Yes', 'No']]
markup = ReplyKeyboardMarkup(reply_keyboard, one_time_keyboard=True)

# init functions
def validate_nric(nric):
    if len(nric) == 5:
        if nric[:4].isdigit() and nric[4].isalpha():
            return True
    return False

###############

def start(bot, update):
    update.message.reply_text(
''' _Welcome to the Redesign Seminar!_
*============================*
I'm here to mark your attendance and provide your seat letter. You may leave this chat at any time, it will not affect the bot.

Please enter the _last 5 characters_ of your NRIC:''',
        parse_mode = 'Markdown')

    logger.info("User %s initiates contact", update.message.from_user.first_name)

    return TYPING_NRIC


def get_nric(bot, update, user_data):
    text = update.message.text
    user_data['NRIC'] = text
    if validate_nric(text):
        update.message.reply_text(
            "To confirm, the last 5 characters of your NRIC are {}.\n\nIs this correct? Yes/No".format(text), reply_markup=markup)

        return RESPONSE
    
    else:
        update.message.reply_text(
            "*The provided partial NRIC {} is incorrect!* Please check that it is in the format 1234E (where full NRIC would be S9951234E)\n\nLet's try again. Last 5 characters of your NRIC:".format(text),
            parse_mode = 'Markdown')
        
        return TYPING_NRIC


def final(bot, update, user_data):
    text = update.message.text
    if text.lower() == "yes":
        #MAIN ACTION
        seating = returnSeating(PERSON, user_data['NRIC'])
        if seating == None:
            update.message.reply_text("We cannot find your NRIC, please look for assistance around the venue.\n\nYou may now leave this chat or type /start to register for another attendee.")
        else:
            update.message.reply_text('''Your seating is: {}.\n\nThank you for attending this seminar! You may now leave this chat or type /start to register for another attendee.'''.format(seating))
    elif text.lower() == "no":
        update.message.reply_text("Let's try again. Enter the last 5 digits of your NRIC:")
        return TYPING_NRIC

    logger.info("User {} completes".format(update.message.from_user.first_name))
    user_data.clear()
    return ConversationHandler.END

###############

# admin stuff
def collectStats(bot, update):
    if update.message.chat_id == 234058962:
        count = 0
        total = 0
        for each in PERSON:
            total += 1
            if each['GRP1_REG'] == 'P':
                count += 1
        update.message.reply_text("Good day, admin.\nTotal: {}, Present: {}".format(total, count))
        logger.info("Admin initiates stats report.\nTotal: {}, Present: {}".format(total, count))
    else:
        update.message.reply_text("You are not recognised!")

def error(bot, update, error):
    """Log Errors caused by Updates."""
    logger.warning('Update "%s" caused error "%s"', update, error)

###############

def main():
    # Create the Updater and pass it your bot's token.
    updater = Updater("241346491:AAH09_cf9KfaFohGgXUo96ljvOeyqcD1k4o")

    # Get the dispatcher to register handlers
    dp = updater.dispatcher

    # Add conversation handler with the states GENDER, PHOTO, LOCATION and BIO
    conv_handler = ConversationHandler(
        entry_points=[CommandHandler('start', start)],
        states={
            TYPING_NRIC: [MessageHandler(Filters.text,
                                           get_nric,
                                           pass_user_data=True)],
            RESPONSE: [RegexHandler('^Yes|No$',
                                          final,
                                          pass_user_data=True)]
        },
        fallbacks=[]
    )

    def killBot(bot, update):
        chat_id = update.message.chat_id
        if update.message.chat_id == 234058962:
            update.message.reply_text("Saving memory to Excel file...")
            saveFile(PERSON, ws, main_workbook)
            bot.send_document(chat_id, document=open('SeminarDatasheet.xlsx', 'rb'))
            update.message.reply_text("Bot is being killed...")
            logger.info("Bot has been killed.")
            updater.stop()
            updater.is_idle = False
        else:
            update.message.reply_text("You are not recognised!")


    dp.add_handler(conv_handler)
    dp.add_handler(CommandHandler('kill', killBot))
    dp.add_handler(CommandHandler('stats', collectStats))

    # log all errors
    dp.add_error_handler(error)

    # Start the Bot
    updater.start_polling()
    logger.info("The bot is running.")
    # Run the bot until you press Ctrl-C or the process receives SIGINT,
    # SIGTERM or SIGABRT. This should be used most of the time, since
    # start_polling() is non-blocking and will stop the bot gracefully.
    updater.idle()


if __name__ == '__main__':
    main()